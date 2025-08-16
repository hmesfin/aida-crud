import csv
import io
import json
from datetime import datetime

import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter


class DataExporter:
    """Handle data export in various formats."""

    @staticmethod
    def export_csv(queryset, fields=None, filename=None):
        """Export queryset to CSV."""
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        if not fields and queryset:
            model = queryset.model
            fields = [f.name for f in model._meta.fields]

        writer = csv.writer(response)
        writer.writerow(fields)

        for obj in queryset:
            row = []
            for field in fields:
                value = getattr(obj, field)
                if hasattr(value, "isoformat"):
                    value = value.isoformat()
                row.append(value)
            writer.writerow(row)

        return response

    @staticmethod
    def export_json(queryset, fields=None, filename=None):
        """Export queryset to JSON."""
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

        data = []

        if not fields and queryset:
            model = queryset.model
            fields = [f.name for f in model._meta.fields]

        for obj in queryset:
            item = {}
            for field in fields:
                value = getattr(obj, field)
                if hasattr(value, "isoformat"):
                    value = value.isoformat()
                elif hasattr(value, "pk"):
                    value = str(value.pk)
                item[field] = value
            data.append(item)

        response = HttpResponse(
            json.dumps(data, indent=2, default=str), content_type="application/json"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    @staticmethod
    def export_excel(queryset, fields=None, filename=None):
        """Export queryset to Excel"""
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Export"

        if not fields and queryset:
            model = queryset.model
            fields = [f.name for f in model._meta.fields]

        for col_num, field_name in enumerate(fields, 1):
            column_letter = get_column_letter(col_num)
            cell = worksheet[f"{column_letter}1"]
            cell.value = field_name
            cell.font = openpyxl.styles.Font(bold=True)
            worksheet.column_dimensions[column_letter].width = 15

        for row_num, obj in enumerate(queryset, 2):
            for col_num, field in enumerate(fields, 1):
                value = getattr(obj, field)
                if hasattr(value, "isoformat"):
                    value = value.isoformat()
                elif hasattr(value, "pk"):
                    value = str(value.pk)
                worksheet.cell(row=row_num, column=col_num, value=value)

        virtual_workbook = io.BytesIO()
        workbook.save(virtual_workbook)
        virtual_workbook.seek(0)

        response = HttpResponse(
            virtual_workbook.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    @classmethod
    def export(cls, queryset, format="csv", fields=None, filename=None):
        """Export data in specified format"""
        exporters = {
            "csv": cls.export_csv,
            "json": cls.export_json,
            "xlsx": cls.export_excel,
            "excel": cls.export_excel,
        }

        exporter = exporters.get(format, cls.export_csv)
        return exporter(queryset, fields, filename)
